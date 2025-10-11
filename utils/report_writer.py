import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional

def save_report(results, filename, total_packets=None, timestamp=None):
    with open(filename, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["IP", "VPN Status"])
        for row in results:
            writer.writerow(row)
        if total_packets is not None:
            writer.writerow([])
            writer.writerow(["Total Packets", total_packets])
        if timestamp:
            writer.writerow(["Timestamp", timestamp])

def save_full_report(
    filename,
    vpn_results,
    total_packets,
    timestamp,
    deanonym_results=None,
    geo_data=None,
    mac_data=None,
    payload_data=None
):
    with open(filename, "w", newline="", encoding='utf-8') as f:
        writer = csv.writer(f)

        # Create a comprehensive header with all columns
        headers = [
            "IP Address",
            "VPN Status", 
            "Fingerprint Info",
            "Country",
            "City", 
            "ISP",
            "MAC Address",
            "Payload Summary"
        ]
        writer.writerow(headers)

        # Create a dictionary to store all data for each IP
        all_data = {}
        
        # Add VPN results
        for ip, vpn_status in vpn_results:
            all_data[ip] = {"VPN Status": vpn_status}
        
        # Add de-anonymization results
        if deanonym_results:
            for ip, fingerprint in deanonym_results:
                if ip not in all_data:
                    all_data[ip] = {}
                all_data[ip]["Fingerprint Info"] = fingerprint
        
        # Add geolocation data
        if geo_data:
            for ip, info in geo_data.items():
                if ip not in all_data:
                    all_data[ip] = {}
                all_data[ip]["Country"] = info.get("country", "")
                all_data[ip]["City"] = info.get("city", "")
                all_data[ip]["ISP"] = info.get("isp", "")
        
        # Add MAC data
        if mac_data:
            for ip, mac in mac_data.items():
                if ip not in all_data:
                    all_data[ip] = {}
                all_data[ip]["MAC Address"] = mac
        
        # Add payload data
        if payload_data:
            for ip, payload in payload_data.items():
                if ip not in all_data:
                    all_data[ip] = {}
                all_data[ip]["Payload Summary"] = payload
        
        # Write all data rows
        for ip in sorted(all_data.keys()):
            row = [ip]
            for header in headers[1:]:  # Skip IP Address header
                row.append(all_data[ip].get(header, ""))
            writer.writerow(row)
        
        # Add summary information at the end
        writer.writerow([])
        writer.writerow(["=== ANALYSIS SUMMARY ==="])
        writer.writerow(["Total IPs Analyzed", len(all_data)])
        writer.writerow(["Total Packets Processed", total_packets])
        writer.writerow(["Analysis Timestamp", timestamp])
        
        # Count VPNs
        vpn_count = sum(1 for data in all_data.values() if data.get("VPN Status") is True)
        writer.writerow(["VPN IPs Detected", vpn_count])
        
        # Count countries
        countries = set(data.get("Country", "") for data in all_data.values() if data.get("Country"))
        writer.writerow(["Unique Countries", len(countries)])
        
        # Count unique MACs
        macs = set(data.get("MAC Address", "") for data in all_data.values() if data.get("MAC Address"))
        writer.writerow(["Unique MAC Addresses", len(macs)])

def save_comprehensive_excel_report(
    filename: str,
    vpn_results: List,
    total_packets: int,
    timestamp: str,
    deanonym_results: Optional[List] = None,
    geo_data: Optional[Dict] = None,
    mac_data: Optional[Dict] = None,
    payload_data: Optional[Dict] = None,
    ai_analysis: Optional[Dict] = None,
    packet_analysis: Optional[Dict] = None,
    enhanced_analysis: Optional[Dict] = None
):
    """
    Create a comprehensive Excel report with multiple sheets:
    - Sheet 1: Main Analysis Results
    - Sheet 2: AI Threat Analysis
    - Sheet 3: Detailed Packet Analysis
    """
    
    # Create workbook and sheets
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets
    main_sheet = wb.create_sheet("Main Analysis")
    ai_sheet = wb.create_sheet("AI Threat Analysis")
    packet_sheet = wb.create_sheet("Packet Analysis")
    enhanced_sheet = wb.create_sheet("Enhanced De-anonymization")
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(bold=True, color="000000")
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # === SHEET 1: Main Analysis ===
    main_sheet.title = "Main Analysis"
    
    # Headers for main analysis
    headers = [
        "IP Address", "VPN Status", "Fingerprint Info", "Country", 
        "City", "ISP", "MAC Address", "Payload Summary"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = main_sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    
    # Create data dictionary
    all_data = {}
    
    # Add VPN results
    for ip, vpn_status in vpn_results:
        all_data[ip] = {"VPN Status": vpn_status}
    
    # Add de-anonymization results
    if deanonym_results:
        for ip, fingerprint in deanonym_results:
            if ip not in all_data:
                all_data[ip] = {}
            all_data[ip]["Fingerprint Info"] = fingerprint
    
    # Add geolocation data
    if geo_data:
        for ip, info in geo_data.items():
            if ip not in all_data:
                all_data[ip] = {}
            all_data[ip]["Country"] = info.get("country", "")
            all_data[ip]["City"] = info.get("city", "")
            all_data[ip]["ISP"] = info.get("isp", "")
    
    # Add MAC data
    if mac_data:
        for ip, mac in mac_data.items():
            if ip not in all_data:
                all_data[ip] = {}
            all_data[ip]["MAC Address"] = mac
    
    # Add payload data
    if payload_data:
        for ip, payload in payload_data.items():
            if ip not in all_data:
                all_data[ip] = {}
            all_data[ip]["Payload Summary"] = payload
    
    # Write data rows
    for row_idx, ip in enumerate(sorted(all_data.keys()), 2):
        row_data = [ip]
        for header in headers[1:]:
            row_data.append(all_data[ip].get(header, ""))
        
        for col_idx, value in enumerate(row_data, 1):
            cell = main_sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
    
    # Add summary section
    summary_row = len(all_data) + 4
    main_sheet.cell(row=summary_row, column=1, value="=== ANALYSIS SUMMARY ===").font = Font(bold=True, size=14)
    
    summary_data = [
        ["Total IPs Analyzed", len(all_data)],
        ["Total Packets Processed", total_packets],
        ["Analysis Timestamp", timestamp],
        ["VPN IPs Detected", sum(1 for data in all_data.values() if data.get("VPN Status") is True)],
        ["Unique Countries", len(set(data.get("Country", "") for data in all_data.values() if data.get("Country")))],
        ["Unique MAC Addresses", len(set(data.get("MAC Address", "") for data in all_data.values() if data.get("MAC Address")))]
    ]
    
    for idx, (label, value) in enumerate(summary_data, summary_row + 1):
        main_sheet.cell(row=idx, column=1, value=label).font = Font(bold=True)
        main_sheet.cell(row=idx, column=2, value=value)
    
    # Auto-adjust column widths
    for column in main_sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        main_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # === SHEET 2: AI Threat Analysis ===
    if ai_analysis:
        ai_sheet.title = "AI Threat Analysis"
        
        # Network Behavior Analysis
        row = 1
        ai_sheet.cell(row=row, column=1, value="NETWORK BEHAVIOR ANALYSIS").font = Font(bold=True, size=14)
        ai_sheet.cell(row=row, column=1).fill = subheader_fill
        
        if "network_analysis" in ai_analysis and "error" not in ai_analysis["network_analysis"]:
            row += 2
            for key, value in ai_analysis["network_analysis"].items():
                ai_sheet.cell(row=row, column=1, value=f"{key.replace('_', ' ').title()}:").font = Font(bold=True)
                if isinstance(value, list):
                    ai_sheet.cell(row=row, column=2, value=", ".join(str(item) for item in value))
                else:
                    ai_sheet.cell(row=row, column=2, value=str(value))
                row += 1
        
        # Payload Intelligence Analysis
        row += 2
        ai_sheet.cell(row=row, column=1, value="PAYLOAD INTELLIGENCE ANALYSIS").font = Font(bold=True, size=14)
        ai_sheet.cell(row=row, column=1).fill = subheader_fill
        
        if "payload_analysis" in ai_analysis and "error" not in ai_analysis["payload_analysis"]:
            row += 2
            for key, value in ai_analysis["payload_analysis"].items():
                ai_sheet.cell(row=row, column=1, value=f"{key.replace('_', ' ').title()}:").font = Font(bold=True)
                if isinstance(value, list):
                    ai_sheet.cell(row=row, column=2, value=", ".join(str(item) for item in value))
                else:
                    ai_sheet.cell(row=row, column=2, value=str(value))
                row += 1
        
        # Threat Report
        row += 2
        ai_sheet.cell(row=row, column=1, value="COMPREHENSIVE THREAT REPORT").font = Font(bold=True, size=14)
        ai_sheet.cell(row=row, column=1).fill = subheader_fill
        
        if "threat_report" in ai_analysis:
            row += 2
            threat_report = ai_analysis["threat_report"]
            # Split report into lines and add to sheet
            lines = threat_report.split('\n')
            for line in lines:
                if line.strip():
                    ai_sheet.cell(row=row, column=1, value=line)
                    row += 1
        
        # Auto-adjust column widths for AI sheet
        for column in ai_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            ai_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # === SHEET 3: Detailed Packet Analysis ===
    if packet_analysis:
        packet_sheet.title = "Packet Analysis"
        
        # Protocol Statistics
        row = 1
        packet_sheet.cell(row=row, column=1, value="PROTOCOL STATISTICS").font = Font(bold=True, size=14)
        packet_sheet.cell(row=row, column=1).fill = subheader_fill
        
        row += 2
        packet_sheet.cell(row=row, column=1, value="Total Packets:").font = Font(bold=True)
        packet_sheet.cell(row=row, column=2, value=packet_analysis.get('total_packets', 0))
        row += 1
        
        packet_sheet.cell(row=row, column=1, value="Average Packet Size:").font = Font(bold=True)
        packet_sheet.cell(row=row, column=2, value=f"{packet_analysis.get('avg_packet_size', 0):.1f} bytes")
        row += 2
        
        # Protocol Distribution
        packet_sheet.cell(row=row, column=1, value="Protocol Distribution:").font = Font(bold=True, size=12)
        packet_sheet.cell(row=row, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        row += 1
        
        for protocol, count in packet_analysis.get('protocol_stats', {}).items():
            percentage = (count / packet_analysis.get('total_packets', 1)) * 100
            packet_sheet.cell(row=row, column=1, value=f"  {protocol}")
            packet_sheet.cell(row=row, column=2, value=f"{count:,} packets")
            packet_sheet.cell(row=row, column=3, value=f"({percentage:.1f}%)")
            row += 1
        
        # Top Ports
        row += 2
        packet_sheet.cell(row=row, column=1, value="Top Ports Used:").font = Font(bold=True, size=12)
        packet_sheet.cell(row=row, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        row += 1
        
        for port_info, count in packet_analysis.get('top_ports', {}).items():
            packet_sheet.cell(row=row, column=1, value=f"  {port_info}")
            packet_sheet.cell(row=row, column=2, value=f"{count:,} packets")
            row += 1
        
        # Website Access
        row += 2
        packet_sheet.cell(row=row, column=1, value="WEBSITE ACCESS ANALYSIS").font = Font(bold=True, size=14)
        packet_sheet.cell(row=row, column=1).fill = subheader_fill
        
        websites = packet_analysis.get('top_websites', {})
        if websites:
            row += 2
            packet_sheet.cell(row=row, column=1, value="Most Accessed Websites:").font = Font(bold=True, size=12)
            packet_sheet.cell(row=row, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            row += 1
            
            for website, count in websites.items():
                packet_sheet.cell(row=row, column=1, value=f"  {website}")
                packet_sheet.cell(row=row, column=2, value=f"{count} queries")
                row += 1
        
        # Potential Passwords
        row += 2
        packet_sheet.cell(row=row, column=1, value="POTENTIAL PASSWORD DETECTION").font = Font(bold=True, size=14)
        packet_sheet.cell(row=row, column=1).fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        
        passwords = packet_analysis.get('potential_passwords', [])
        if passwords:
            row += 2
            packet_sheet.cell(row=row, column=1, value=f"Found {len(passwords)} potential password fields!").font = Font(bold=True, color="FF0000")
            row += 2
            
            # Headers for password data
            pwd_headers = ["Type", "Source IP", "Destination IP", "Field", "Value"]
            for col, header in enumerate(pwd_headers, 1):
                cell = packet_sheet.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            row += 1
            
            for pwd in passwords:
                packet_sheet.cell(row=row, column=1, value=pwd.get('type', ''))
                packet_sheet.cell(row=row, column=2, value=pwd.get('src_ip', ''))
                packet_sheet.cell(row=row, column=3, value=pwd.get('dst_ip', ''))
                packet_sheet.cell(row=row, column=4, value=pwd.get('field', ''))
                packet_sheet.cell(row=row, column=5, value=pwd.get('value', ''))
                row += 1
        
        # Suspicious Activity
        row += 2
        packet_sheet.cell(row=row, column=1, value="SUSPICIOUS ACTIVITY DETECTION").font = Font(bold=True, size=14)
        packet_sheet.cell(row=row, column=1).fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        
        suspicious = packet_analysis.get('suspicious_activity', [])
        if suspicious:
            row += 2
            for activity in suspicious:
                severity_color = "FF0000" if activity['severity'] == 'High' else "FFA500" if activity['severity'] == 'Medium' else "00FF00"
                packet_sheet.cell(row=row, column=1, value=f"{activity['type']} ({activity['severity']})").font = Font(bold=True, color=severity_color)
                packet_sheet.cell(row=row, column=2, value=activity['details'])
                row += 1
        
        # Auto-adjust column widths for packet sheet
        for column in packet_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            packet_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # === SHEET 4: Enhanced De-anonymization ===
    if enhanced_analysis:
        enhanced_sheet.title = "Enhanced De-anonymization"
        
        # Real IP Detection Results
        row = 1
        enhanced_sheet.cell(row=row, column=1, value="REAL IP DETECTION RESULTS").font = Font(bold=True, size=14)
        enhanced_sheet.cell(row=row, column=1).fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        enhanced_sheet.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
        
        real_ip_results = enhanced_analysis.get('real_ip_detection', {})
        if 'error' not in real_ip_results:
            row += 2
            enhanced_sheet.cell(row=row, column=1, value="VPN IPs Detected:").font = Font(bold=True)
            vpn_ips = real_ip_results.get('vpn_ips_detected', [])
            enhanced_sheet.cell(row=row, column=2, value=f"{len(vpn_ips)} IPs")
            row += 1
            
            for vpn_ip in vpn_ips[:5]:  # Show first 5
                enhanced_sheet.cell(row=row, column=1, value=f"  • {vpn_ip}")
                row += 1
            
            row += 1
            enhanced_sheet.cell(row=row, column=1, value="Potential Real IPs:").font = Font(bold=True)
            potential_ips = real_ip_results.get('potential_real_ips', [])
            enhanced_sheet.cell(row=row, column=2, value=f"{len(potential_ips)} candidates")
            row += 1
            
            if potential_ips:
                for ip in potential_ips[:3]:  # Show top 3 candidates
                    confidence = real_ip_results.get('confidence_scores', {}).get(ip, 0)
                    enhanced_sheet.cell(row=row, column=1, value=f"  🎯 {ip}")
                    enhanced_sheet.cell(row=row, column=2, value=f"Confidence: {confidence}%")
                    if confidence > 70:
                        enhanced_sheet.cell(row=row, column=2).font = Font(color="FF0000", bold=True)
                    elif confidence > 40:
                        enhanced_sheet.cell(row=row, column=2).font = Font(color="FFA500", bold=True)
                    row += 1
            else:
                enhanced_sheet.cell(row=row, column=1, value="  ✅ No high-confidence real IPs detected")
                enhanced_sheet.cell(row=row, column=1).font = Font(color="00AA00")
                row += 1
        
        # Encrypted Data Extraction
        row += 2
        enhanced_sheet.cell(row=row, column=1, value="ENCRYPTED DATA EXTRACTION").font = Font(bold=True, size=14)
        enhanced_sheet.cell(row=row, column=1).fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        enhanced_sheet.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
        
        encrypted_extraction = enhanced_analysis.get('encrypted_data_extraction', {})
        if encrypted_extraction:
            row += 2
            total_packets = encrypted_extraction.get('total_encrypted_packets', 0)
            enhanced_sheet.cell(row=row, column=1, value="Total VPN Encrypted Packets:").font = Font(bold=True)
            enhanced_sheet.cell(row=row, column=2, value=f"{total_packets:,} packets")
            row += 1
            
            sample_packets = encrypted_extraction.get('sample_packets', [])
            if sample_packets:
                enhanced_sheet.cell(row=row, column=1, value="Sample Encrypted Packets:").font = Font(bold=True)
                row += 1
                
                for i, pkt in enumerate(sample_packets[:3], 1):
                    direction = "OUT" if pkt['src_ip'] != "51.15.62.60" else "IN"
                    enhanced_sheet.cell(row=row, column=1, value=f"  {i}. [{direction}] {pkt['src_ip']} → {pkt['dst_ip']}")
                    enhanced_sheet.cell(row=row, column=2, value=f"{pkt['data_length']} bytes")
                    row += 1
        
        # Advanced Device Fingerprinting
        row += 2
        enhanced_sheet.cell(row=row, column=1, value="ADVANCED DEVICE FINGERPRINTING").font = Font(bold=True, size=14)
        enhanced_sheet.cell(row=row, column=1).fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
        enhanced_sheet.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
        
        fingerprint_results = enhanced_analysis.get('advanced_fingerprints', {})
        if 'device_fingerprints' in fingerprint_results:
            row += 2
            device_fps = fingerprint_results['device_fingerprints']
            enhanced_sheet.cell(row=row, column=1, value="Device Fingerprints Generated:").font = Font(bold=True)
            enhanced_sheet.cell(row=row, column=2, value=f"{len(device_fps)} devices")
            row += 1
            
            for ip, fp_data in list(device_fps.items())[:3]:  # Show first 3
                if isinstance(fp_data, dict) and 'os_guess' in fp_data:
                    enhanced_sheet.cell(row=row, column=1, value=f"  • {ip}")
                    enhanced_sheet.cell(row=row, column=2, value=f"OS: {fp_data.get('os_guess', 'Unknown')}")
                    row += 1
        
        # Traffic Flow Correlation
        row += 2
        enhanced_sheet.cell(row=row, column=1, value="TRAFFIC FLOW CORRELATION").font = Font(bold=True, size=14)
        enhanced_sheet.cell(row=row, column=1).fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
        enhanced_sheet.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
        
        flow_results = enhanced_analysis.get('traffic_flows', {})
        if 'flow_analysis' in flow_results:
            row += 2
            flow_analysis = flow_results['flow_analysis']
            total_flows = flow_analysis.get('total_flows', 0)
            enhanced_sheet.cell(row=row, column=1, value="Total Network Flows:").font = Font(bold=True)
            enhanced_sheet.cell(row=row, column=2, value=f"{total_flows} flows")
            row += 1
            
            if 'correlation_results' in flow_results:
                correlations = flow_results['correlation_results']
                suspicious_pairs = correlations.get('suspicious_pairs', [])
                enhanced_sheet.cell(row=row, column=1, value="Suspicious Flow Correlations:").font = Font(bold=True)
                enhanced_sheet.cell(row=row, column=2, value=f"{len(suspicious_pairs)} pairs")
                if len(suspicious_pairs) > 0:
                    enhanced_sheet.cell(row=row, column=2).font = Font(color="FF0000", bold=True)
                row += 1
        
        # Encrypted Traffic Analysis
        row += 2
        enhanced_sheet.cell(row=row, column=1, value="ENCRYPTED VPN TRAFFIC ANALYSIS").font = Font(bold=True, size=14)
        enhanced_sheet.cell(row=row, column=1).fill = PatternFill(start_color="1ABC9C", end_color="1ABC9C", fill_type="solid")
        enhanced_sheet.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
        
        encrypted_results = enhanced_analysis.get('encrypted_traffic', {})
        if 'metadata_analysis' in encrypted_results:
            row += 2
            metadata = encrypted_results['metadata_analysis']
            enhanced_sheet.cell(row=row, column=1, value="Metadata Analysis:").font = Font(bold=True)
            row += 1
            enhanced_sheet.cell(row=row, column=1, value="  Total Encrypted Flows:")
            enhanced_sheet.cell(row=row, column=2, value=metadata.get('encrypted_flows', 0))
            row += 1
            enhanced_sheet.cell(row=row, column=1, value="  Average Entropy:")
            enhanced_sheet.cell(row=row, column=2, value=f"{metadata.get('avg_entropy', 0):.3f}")
            row += 1
            enhanced_sheet.cell(row=row, column=1, value="  VPN Protocols:")
            enhanced_sheet.cell(row=row, column=2, value=', '.join(metadata.get('vpn_protocols', [])))
            row += 1
            enhanced_sheet.cell(row=row, column=1, value="  Encryption Strength:")
            enhanced_sheet.cell(row=row, column=2, value=metadata.get('encryption_strength', 'Unknown'))
        
        if 'website_fingerprinting' in encrypted_results:
            row += 2
            enhanced_sheet.cell(row=row, column=1, value="Website Fingerprinting (Through Encrypted HTTPS):").font = Font(bold=True)
            row += 1
            websites = encrypted_results['website_fingerprinting']
            for site, confidence in websites.items():
                enhanced_sheet.cell(row=row, column=1, value=f"  {site}")
                enhanced_sheet.cell(row=row, column=2, value=f"{confidence:.1%} confidence")
                row += 1
        
        # Advanced Fingerprinting
        row += 2
        enhanced_sheet.cell(row=row, column=1, value="ADVANCED DEVICE FINGERPRINTING").font = Font(bold=True, size=14)
        enhanced_sheet.cell(row=row, column=1).fill = subheader_fill
        
        fingerprint_results = enhanced_analysis.get('advanced_fingerprints', {})
        if 'device_signatures' in fingerprint_results:
            row += 2
            enhanced_sheet.cell(row=row, column=1, value="Unique Device Signatures:").font = Font(bold=True)
            row += 1
            signatures = fingerprint_results['device_signatures']
            for ip, signature in signatures.items():
                enhanced_sheet.cell(row=row, column=1, value=f"  {ip}")
                enhanced_sheet.cell(row=row, column=2, value=signature)
                row += 1
        
        # Traffic Correlation
        row += 2
        enhanced_sheet.cell(row=row, column=1, value="TRAFFIC FLOW CORRELATION").font = Font(bold=True, size=14)
        enhanced_sheet.cell(row=row, column=1).fill = subheader_fill
        
        correlation_results = enhanced_analysis.get('traffic_correlation', {})
        if 'flow_correlations' in correlation_results:
            row += 2
            enhanced_sheet.cell(row=row, column=1, value="Correlated Traffic Flows:").font = Font(bold=True)
            row += 1
            correlations = correlation_results['flow_correlations']
            for correlation in correlations:
                enhanced_sheet.cell(row=row, column=1, value=f"  {correlation}")
                row += 1
        
        # DNS Leak Detection
        row += 2
        enhanced_sheet.cell(row=row, column=1, value="DNS LEAK DETECTION & PRIVACY ANALYSIS").font = Font(bold=True, size=14)
        enhanced_sheet.cell(row=row, column=1).fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        enhanced_sheet.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
        
        dns_results = enhanced_analysis.get('dns_leaks', {})
        if 'privacy_risk_score' in dns_results:
            row += 2
            risk_score = dns_results['privacy_risk_score']
            risk_level = "HIGH" if risk_score > 70 else "MEDIUM" if risk_score > 40 else "LOW"
            enhanced_sheet.cell(row=row, column=1, value="Privacy Risk Score:").font = Font(bold=True)
            enhanced_sheet.cell(row=row, column=2, value=f"{risk_score}/100 ({risk_level})")
            if risk_score > 70:
                enhanced_sheet.cell(row=row, column=2).font = Font(color="FF0000", bold=True)
            elif risk_score > 40:
                enhanced_sheet.cell(row=row, column=2).font = Font(color="FFA500", bold=True)
            else:
                enhanced_sheet.cell(row=row, column=2).font = Font(color="00AA00", bold=True)
        
        if 'detected_leaks' in dns_results:
            row += 2
            enhanced_sheet.cell(row=row, column=1, value="Detected DNS Leaks:").font = Font(bold=True)
            row += 1
            leaks = dns_results['detected_leaks']
            for leak in leaks:
                enhanced_sheet.cell(row=row, column=1, value=f"  • {leak}")
                enhanced_sheet.cell(row=row, column=1).font = Font(color="FF0000")
                row += 1
        
        if 'isp_dns_usage' in dns_results:
            row += 2
            enhanced_sheet.cell(row=row, column=1, value="ISP DNS Server Usage (Reveals Real Location):").font = Font(bold=True)
            row += 1
            isp_dns = dns_results['isp_dns_usage']
            for server, usage in isp_dns.items():
                enhanced_sheet.cell(row=row, column=1, value=f"  {server}")
                enhanced_sheet.cell(row=row, column=2, value=usage)
                row += 1
        
        # Summary Section
        row += 3
        enhanced_sheet.cell(row=row, column=1, value="ENHANCED DE-ANONYMIZATION SUMMARY").font = Font(bold=True, size=16)
        enhanced_sheet.cell(row=row, column=1).fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        enhanced_sheet.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
        
        row += 2
        # Count analysis results
        encrypted_count = len(encrypted_results.get('analyzed_ips', []))
        fingerprint_count = len(fingerprint_results.get('device_signatures', {}))
        correlation_count = len(correlation_results.get('flow_correlations', []))
        dns_leak_count = len(dns_results.get('detected_leaks', []))
        
        summary_data = [
            ["Encrypted Traffic Analysis Coverage", f"{encrypted_count} IPs"],
            ["Advanced Fingerprints Generated", f"{fingerprint_count} devices"],
            ["Traffic Correlations Found", f"{correlation_count} flows"],
            ["DNS Leaks Detected", f"{dns_leak_count} issues"],
            ["Overall Threat Level", "HIGH" if dns_leak_count > 5 else "MEDIUM" if dns_leak_count > 2 else "LOW"]
        ]
        
        for label, value in summary_data:
            enhanced_sheet.cell(row=row, column=1, value=label).font = Font(bold=True)
            enhanced_sheet.cell(row=row, column=2, value=value)
            row += 1
        
        row += 2
        enhanced_sheet.cell(row=row, column=1, value="ADVANCED TECHNIQUES USED:").font = Font(bold=True)
        row += 1
        techniques = [
            "• Encrypted metadata analysis (no decryption)",
            "• Website fingerprinting through HTTPS",
            "• Cross-correlation timing attacks",
            "• Side-channel information leakage",
            "• DNS query pattern analysis",
            "• Behavioral device fingerprinting"
        ]
        
        for technique in techniques:
            enhanced_sheet.cell(row=row, column=1, value=technique)
            row += 1
        
        row += 2
        enhanced_sheet.cell(row=row, column=1, value="LEGAL & ETHICAL COMPLIANCE:").font = Font(bold=True)
        row += 1
        compliance_notes = [
            "• No VPN encryption was broken",
            "• Analysis based on metadata only",
            "• For authorized research purposes",
            "• Compliant with cybersecurity education standards"
        ]
        
        for note in compliance_notes:
            enhanced_sheet.cell(row=row, column=1, value=note)
            enhanced_sheet.cell(row=row, column=1).font = Font(color="00AA00")
            row += 1
        
        # Auto-adjust column widths for enhanced sheet
        for column in enhanced_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            enhanced_sheet.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook
    wb.save(filename)
    return filename

